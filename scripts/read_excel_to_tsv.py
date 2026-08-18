#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
读取税务局导出的Excel发票记录，按开票人规则筛选并输出18列TSV（用于粘贴到企微表格）。

用法：
    python read_excel_to_tsv.py <excel文件路径>

输出：
    stdout：18列TSV的发票数据（仅含 符瑞香/伍惠娟 两类开票人的记录）。
    stderr：统计信息 + 异常开票人明细（若有）。

开票人规则（V2.0 核心）：
    符瑞香  → 正常录入，是否邮件开票留空
    伍惠娟  → 正常录入，是否邮件开票 = "是"
    其他    → 异常，不录入，转人工判断（脚本打印明细并以退出码 2 结束）

18列TSV列布局（对应企微表格引擎列，0-based）：
    col0  ""              空
    col1  ""              公司(留空)
    col2  开票日期        ← Excel col9  (idx8)
    col3  ""              发票代码(留空)
    col4  发票号码        ← Excel col4  (idx3)
    col5  发票类型        ← Excel col22 (idx21)
    col6  开票名称        ← Excel col8  (idx7)
    col7  纳税人识别号    ← Excel col7  (idx6)
    col8  开票金额        ← Excel col20 (idx19)
    col9  订单ID          ← Excel col27 (idx26)
    col10 ""              备注(留空)
    col11 ""              归属项目(留空)
    col12 ""              备注_1(留空)
    col13 ""              订单ID是否重复(留空)
    col14 ""              年(留空)
    col15 ""              月(留空)
    col16 ""              货物或应税劳务名称(留空)
    col17 是否邮件开票    ← 伍惠娟="是"，其余=""
"""
import sys
import openpyxl
from datetime import datetime

# 开票人 → 是否邮件开票 映射规则（业务约定，改动只改这里）
NORMAL_ISSUERS = {"符瑞香"}   # 直接录入，邮件开票留空
MAIL_ISSUERS = {"伍惠娟"}     # 录入 + 邮件开票 = "是"
# 不在上述两个集合里的开票人 = 异常，转人工


def format_date(val):
    """格式化日期：2026-07-24 18:22:02 → 2026/7/24"""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return f"{val.year}/{val.month}/{val.day}"
    s = str(val).strip()
    if not s:
        return ""
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y/%m/%d %H:%M:%S"]:
        try:
            dt = datetime.strptime(s, fmt)
            return f"{dt.year}/{dt.month}/{dt.day}"
        except ValueError:
            continue
    return s


def cell_str(val):
    """转字符串，None→空；整型浮点去小数；去换行"""
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    s = str(val)
    s = s.replace("\n", " ").replace("\r", " ").strip()
    return s


def main():
    if len(sys.argv) < 2:
        print("用法: python read_excel_to_tsv.py <excel文件路径>", file=sys.stderr)
        sys.exit(1)

    filepath = sys.argv[1]
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"错误: 无法打开Excel文件: {e}", file=sys.stderr)
        sys.exit(1)

    # 优先用「信息汇总表」，没有则用第一个sheet
    sheet_name = None
    for sn in wb.sheetnames:
        if '信息汇总' in sn:
            sheet_name = sn
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
        print(f"提示: 未找到'信息汇总表'，使用'{sheet_name}'", file=sys.stderr)

    ws = wb[sheet_name]

    lines = []          # 待录入的18列TSV行
    abnormal = []       # 异常开票人记录（转人工）
    skipped = 0
    normal_count = 0    # 符瑞香
    mail_count = 0      # 伍惠娟
    other_count = {}    # 异常开票人 -> 条数

    for row in ws.iter_rows(min_row=2, values_only=True):
        # 跳过空行
        if not row or all(v is None for v in row):
            skipped += 1
            continue

        # 跳过合计行（第一列是"合计"或类似）
        first_cell = str(row[0]) if row[0] else ""
        if "合计" in first_cell:
            skipped += 1
            continue

        # 确保有足够列（信息汇总表共27列，开票人=col26/idx25，备注=col27/idx26）
        if len(row) < 27:
            skipped += 1
            continue

        # 提取字段
        date = format_date(row[8])        # col 9 开票日期
        invoice_num = cell_str(row[3])    # col 4 数电发票号码
        invoice_type = cell_str(row[21])  # col 22 发票票种
        name = cell_str(row[7])           # col 8 购买方名称
        tax_id = cell_str(row[6])         # col 7 购方识别号
        amount = cell_str(row[19])        # col 20 价税合计
        order_id = cell_str(row[26])      # col 27 备注
        issuer = cell_str(row[25])        # col 26 开票人 ← V2.0 新增

        # 跳过没有发票号码的行
        if not invoice_num:
            skipped += 1
            continue

        # ---- V2.0 开票人规则分流 ----
        if issuer in MAIL_ISSUERS:
            mail_flag = "是"
            mail_count += 1
        elif issuer in NORMAL_ISSUERS:
            mail_flag = ""
            normal_count += 1
        else:
            # 异常开票人：不录入，转人工
            abnormal.append({
                "发票号码": invoice_num,
                "开票人": issuer,
                "开票日期": date,
                "订单ID": order_id,
                "开票名称": name,
            })
            other_count[issuer] = other_count.get(issuer, 0) + 1
            continue

        # 组装18列TSV行（制表符分隔）
        tsv_line = "\t".join([
            "",          # col0  空
            "",          # col1  公司(留空)
            date,        # col2  开票日期
            "",          # col3  发票代码(留空)
            invoice_num, # col4  发票号码
            invoice_type, # col5 发票类型
            name,        # col6  开票名称
            tax_id,      # col7  纳税人识别号
            amount,      # col8  开票金额
            order_id,    # col9  订单ID
            "",          # col10 备注(留空)
            "",          # col11 归属项目(留空)
            "",          # col12 备注_1(留空)
            "",          # col13 订单ID是否重复(留空)
            "",          # col14 年(留空)
            "",          # col15 月(留空)
            "",          # col16 货物或应税劳务名称(留空)
            mail_flag,   # col17 是否邮件开票
        ])
        lines.append(tsv_line)

    wb.close()

    # 输出TSV到stdout
    print("\n".join(lines))

    # 统计信息到stderr
    print(f"\n# 共 {len(lines)} 条待录入，跳过 {skipped} 行", file=sys.stderr)
    print(f"#   符瑞香（直接录入）: {normal_count} 条", file=sys.stderr)
    print(f"#   伍惠娟（邮件开票=是）: {mail_count} 条", file=sys.stderr)

    if abnormal:
        print("\n" + "=" * 60, file=sys.stderr)
        print(f"⚠️  发现 {len(abnormal)} 条异常开票人记录，需转人工判断！", file=sys.stderr)
        print("=" * 60, file=sys.stderr)
        for k, v in other_count.items():
            print(f"   开票人「{k}」: {v} 条", file=sys.stderr)
        print("\n异常明细（不录入，转人工）：", file=sys.stderr)
        for a in abnormal:
            print(f"   - 发票号码 {a['发票号码']} | 开票人 {a['开票人']} | 日期 {a['开票日期']} | 订单 {a['订单ID']}", file=sys.stderr)
        print("\n请人工判断这些记录是否录入、是否邮件开票。", file=sys.stderr)
        sys.exit(2)  # 异常信号：提示调用方有需人工处理的记录


if __name__ == "__main__":
    main()
